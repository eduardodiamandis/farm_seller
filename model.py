# model.py
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import datetime
from io import BytesIO

def expandir_coluna_e_salvar_v3(
    caminho_entrada,
    aba,
    coluna_index_data,
    coluna_index_week_number,
    repeticoes,
    lista_estados,
    intervalo_valor=("C", "R"),
    intervalo_percent=("U", "AJ"),
    linha_base_valor=6,
    linha_base_percent=6,
    num_output_rows=None,
    nome_aba=None
):
    """
    Expande dados de uma planilha Excel e retorna um buffer com o arquivo processado.
    
    Parâmetros:
        caminho_entrada (str): Caminho do arquivo Excel de entrada
        aba (str): Nome da aba a ser processada
        coluna_index_data (int): Índice da coluna de datas
        coluna_index_week_number (int): Índice da coluna de números de semana
        repeticoes (int): Número de repetições por data
        lista_estados (list): Lista de estados para preenchimento cíclico
        intervalo_valor (tuple): Intervalo de colunas para fórmulas de valor
        intervalo_percent (tuple): Intervalo de colunas para fórmulas de percentual
        linha_base_valor (int): Linha base para as fórmulas de valor
        linha_base_percent (int): Linha base para as fórmulas de percentual
        num_output_rows (int, optional): Número total de linhas de saída
        nome_aba (str, optional): Nome da aba para referência nas fórmulas
        
    Retorna:
        BytesIO: Buffer com o arquivo Excel processado
        int: Número total de linhas geradas
    """
    # Carrega a planilha original
    wb_original = load_workbook(caminho_entrada)
    ws = wb_original[aba]

    # Coleta os valores da coluna de datas
    original_dates = [
        cell[0].value
        for cell in ws.iter_rows(min_col=coluna_index_data, max_col=coluna_index_data)
        if cell[0].value is not None
    ]
    
    # Coleta os valores da coluna de week numbers
    original_week_numbers = [
        cell[0].value
        for cell in ws.iter_rows(min_col=coluna_index_week_number, max_col=coluna_index_week_number)
        if cell[0].value is not None
    ]

    # Define o número total de linhas
    total_rows = len(original_dates) * repeticoes if num_output_rows is None else num_output_rows

    # Cria nova planilha
    wb_novo = Workbook()
    ws_novo = wb_novo.active
    ws_novo.title = "dados_expandido"

    # Cabeçalhos fixos
    headers = ["WEEK_NUMBER", "YEAR", "DATA", "COUNTRY", "STATES", "VALOR", "TIPO", "PERCENT"]
    for idx, header in enumerate(headers, start=1):
        ws_novo.cell(row=1, column=idx, value=header)

    # Colunas para fórmulas VALOR
    col_inicio_valor = column_index_from_string(intervalo_valor[0])
    col_fim_valor = column_index_from_string(intervalo_valor[1])
    colunas_valor = [get_column_letter(i) for i in range(col_inicio_valor, col_fim_valor + 1)]
    num_cols_valor = len(colunas_valor)

    # Colunas para fórmulas PERCENT
    col_inicio_percent = column_index_from_string(intervalo_percent[0])
    col_fim_percent = column_index_from_string(intervalo_percent[1])
    colunas_percent = [get_column_letter(i) for i in range(col_inicio_percent, col_fim_percent + 1)]
    num_cols_percent = len(colunas_percent)

    for i in range(total_rows):
        row_idx = i + 2  # Começa na linha 2

        # DATA (Read from file and repeat)
        date_index = i // repeticoes
        data = original_dates[date_index % len(original_dates)]
        ws_novo.cell(row=row_idx, column=3, value=data)

        # WEEKNUM (Read from file and repeat)
        if date_index < len(original_week_numbers):
             week_number_val = original_week_numbers[date_index]
        else:
            week_number_val = original_week_numbers[date_index % len(original_week_numbers)]
        ws_novo.cell(row=row_idx, column=1, value=week_number_val)

        # YEAR
        if isinstance(data, datetime.datetime):
            ws_novo.cell(row=row_idx, column=2, value=data.year)
        else:
            try:
                date_obj = datetime.datetime.strptime(data, '%Y-%m-%d %H:%M:%S')
                ws_novo.cell(row=row_idx, column=2, value=date_obj.year)
            except:
                ws_novo.cell(row=row_idx, column=2, value="")

        # COUNTRY fixo
        ws_novo.cell(row=row_idx, column=4, value="BRAZIL")

        # STATES cíclico
        estado = lista_estados[i % len(lista_estados)]
        ws_novo.cell(row=row_idx, column=5, value=estado)

        # VALOR (fórmula externa)
        block_valor = i // num_cols_valor
        linha_valor = linha_base_valor + block_valor
        pos_valor = i % num_cols_valor
        col_valor = colunas_valor[pos_valor]
        formula_valor = f"=[SBS_Regional_Farmer_Selling_Estimates.xlsx]{nome_aba}!${col_valor}${linha_valor}"
        ws_novo.cell(row=row_idx, column=6, value=formula_valor)

        # TIPO fixo
        ws_novo.cell(row=row_idx, column=7, value="KMT")

        # PERCENT (fórmula externa)
        block_percent = i // num_cols_percent
        linha_percent = linha_base_percent + block_percent
        pos_percent = i % num_cols_percent
        col_percent = colunas_percent[pos_percent]
        formula_percent = f"=[SBS_Regional_Farmer_Selling_Estimates.xlsx]{nome_aba}!${col_percent}${linha_percent}"
        ws_novo.cell(row=row_idx, column=8, value=formula_percent)

    # Salva em buffer
    buffer = BytesIO()
    wb_novo.save(buffer)
    buffer.seek(0)
    return buffer, total_rows